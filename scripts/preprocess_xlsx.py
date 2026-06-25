#!/usr/bin/env python3
"""
Excel/CSV Preprocessing Pipeline for SE Requirements Decomposition.

Converts spreadsheets (BOM, pin lists, register maps, electrical specs, test matrices)
into structured Markdown tables for downstream consumption by requirements-decompose.

Usage:
    python preprocess_xlsx.py --input BOM.xlsx --output ./preprocessed/
    python preprocess_xlsx.py --input ./docs/inputs/ --output ./preprocessed/ --batch
    python preprocess_xlsx.py --input data.csv --output ./preprocessed/
"""

import argparse
import json
import logging
import os
import sys
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

logger = logging.getLogger("xlsx_preprocess")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".xlsm", ".csv", ".tsv"}


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class SheetInfo:
    """Metadata for one sheet."""
    name: str
    index: int
    rows: int
    cols: int
    markdown_path: str
    has_header: bool = True
    empty_rows_skipped: int = 0


@dataclass
class SpreadsheetManifest:
    """Output manifest for one spreadsheet file."""
    source: str
    total_sheets: int
    sheets_exported: int
    total_rows: int
    total_cells: int
    sheets: list[SheetInfo] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Sheet reading
# ---------------------------------------------------------------------------

def _read_csv(filepath: str, sep: str = ",") -> list[list[str | None]]:
    """Read CSV/TSV and return as 2D list."""
    import csv

    rows: list[list[str | None]] = []
    with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=sep)
        for row in reader:
            rows.append([cell.strip() if cell else None for cell in row])
    return rows


def _read_xlsx_sheets(
    filepath: str,
    *,
    skip_empty: bool = True,
    max_rows_per_sheet: int = 5000,
) -> tuple[dict[str, list[list[str | None]]], list[str]]:
    """
    Read all sheets from an xlsx/xls file.
    Returns (sheets_dict, issues_list).
    Tries pandas first (faster), falls back to openpyxl for complex files.
    """
    issues: list[str] = []
    sheets: dict[str, list[list[str | None]]] = {}
    ext = Path(filepath).suffix.lower()

    # CSV/TSV — single "sheet"
    if ext in (".csv", ".tsv"):
        try:
            sep = "\t" if ext == ".tsv" else ","
            rows = _read_csv(filepath, sep)
            if not rows:
                issues.append("CSV file is empty")
                return sheets, issues
            sheets["Data"] = rows
            logger.info(f"CSV: {len(rows)} rows × {len(rows[0]) if rows else 0} cols")
        except Exception as exc:
            issues.append(f"CSV read failed: {exc}")
        return sheets, issues

    # Try pandas first
    try:
        import pandas as pd

        xl = pd.ExcelFile(filepath)
        for sheet_name in xl.sheet_names:
            df = pd.read_excel(
                xl, sheet_name=sheet_name,
                dtype=str,  # keep everything as string to avoid inference
                header=None,  # we handle headers ourselves
            ).fillna("")
            if skip_empty and df.empty:
                continue
            if len(df) > max_rows_per_sheet:
                issues.append(
                    f"Sheet '{sheet_name}' has {len(df)} rows — "
                    f"truncated to {max_rows_per_sheet}"
                )
                df = df.head(max_rows_per_sheet)

            rows: list[list[str | None]] = []
            for _, row in df.iterrows():
                rows.append([
                    str(cell).strip() if str(cell).strip() else None
                    for cell in row
                ])
            sheets[sheet_name] = rows
            logger.info(
                f"Sheet '{sheet_name}': {len(rows)} rows × "
                f"{len(rows[0]) if rows else 0} cols"
            )
        return sheets, issues

    except ImportError:
        logger.debug("pandas not installed, trying openpyxl...")
    except Exception as exc:
        issues.append(f"pandas read failed ({exc}), trying openpyxl...")

    # Fallback: openpyxl
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filepath, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            empty_count = 0
            for row in ws.iter_rows(values_only=True):
                cells = [
                    str(cell).strip() if cell is not None else None
                    for cell in row
                ]
                # skip fully empty rows
                if all(c is None for c in cells):
                    empty_count += 1
                    continue
                rows.append(cells)

            if skip_empty and not rows:
                continue

            sheets[sheet_name] = rows
            if empty_count:
                logger.info(
                    f"Sheet '{sheet_name}': {len(rows)} rows, "
                    f"skipped {empty_count} empty"
                )
        wb.close()
        logger.info(f"openpyxl: {len(sheets)} sheet(s) from {filepath}")

    except ImportError:
        issues.append("Neither pandas nor openpyxl installed. Run: pip install pandas openpyxl")
    except Exception as exc:
        issues.append(f"openpyxl read failed: {exc}")

    return sheets, issues


# ---------------------------------------------------------------------------
# Sheet → Markdown
# ---------------------------------------------------------------------------

def _guess_header_index(rows: list[list[str | None]]) -> int:
    """
    Guess which row is the header.
    Returns the index of the most likely header row, or -1 if none found.
    """
    if not rows:
        return -1

    # Simple heuristic: first row with all non-None cells
    for i, row in enumerate(rows[:10]):  # only look at first 10 rows
        if row and all(c is not None for c in row):
            return i
    return 0  # default to first row


def _forward_fill(rows: list[list[str | None]], start_row: int = 1) -> list[list[str | None]]:
    """
    Forward-fill empty cells from the last non-empty value in the same column.
    Handles merged cells in spreadsheets — e.g., Category column where only
    the first row of a group has a value.

    start_row: rows before this index are left untouched (typically the header row).
    """
    if not rows or not rows[0]:
        return rows

    max_cols = max(len(r) for r in rows)
    # Start with empty fill_values — header row should NOT seed data values
    fill_values: list[str | None] = [None] * max_cols

    result: list[list[str | None]] = []
    for i, row in enumerate(rows):
        if i < start_row:
            result.append(list(row))
            continue
        new_row: list[str | None] = []
        for j in range(max_cols):
            cell = row[j] if j < len(row) else None
            cell_str = str(cell).strip() if cell is not None else ""
            if cell_str:
                fill_values[j] = cell_str
                new_row.append(cell_str)
            elif fill_values[j] is not None:
                new_row.append(fill_values[j])
            else:
                new_row.append(None)
        result.append(new_row)

    return result


def _sheet_to_markdown(
    rows: list[list[str | None]],
    sheet_name: str,
) -> tuple[str, SheetInfo]:
    """Convert a sheet's rows into a GitHub-flavored Markdown table."""
    if not rows:
        return f"# {sheet_name}\n\n*(empty sheet)*\n", SheetInfo(
            name=sheet_name, index=0, rows=0, cols=0, markdown_path=""
        )

    # Normalize column count (pad short rows, warn about merged cells)
    max_cols = max(len(r) for r in rows)
    normalized: list[list[str]] = []
    for row in rows:
        padded = [(row[i] or "") if i < len(row) else "" for i in range(max_cols)]
        normalized.append(padded)

    # Trim trailing empty columns (common CSV artifact)
    while max_cols > 1 and all(r[-1] == "" for r in normalized):
        for r in normalized:
            r.pop()
        max_cols -= 1

    # Guess header
    header_idx = _guess_header_index(rows)

    # Build markdown
    lines: list[str] = []
    lines.append(f"## {sheet_name}\n")

    for i, row in enumerate(normalized):
        row_str = "| " + " | ".join(row) + " |"
        lines.append(row_str)

        # Add separator after header row
        if i == header_idx:
            lines.append("| " + " | ".join(["---"] * max_cols) + " |")

    # Auto-link common SE patterns in cells
    md = "\n".join(lines) + "\n"

    info = SheetInfo(
        name=sheet_name,
        index=0,
        rows=len(rows),
        cols=max_cols,
        markdown_path="",
        has_header=(header_idx >= 0),
    )

    return md, info


# ---------------------------------------------------------------------------
# Quality verification
# ---------------------------------------------------------------------------

def verify_xlsx_quality(
    sheets: dict[str, list[list[str | None]]],
    sheet_infos: list[SheetInfo],
    issues: list[str],
) -> dict[str, Any]:
    """Assess spreadsheet extraction quality and return a structured report."""
    checks: list[dict[str, Any]] = []
    scores: list[float] = []

    for info in sheet_infos:
        score = 100.0

        # Check 1: Empty rows ratio
        rows = sheets.get(info.name, [])
        empty_rows = sum(1 for r in rows if all(c is None or str(c).strip() == "" for c in r))
        if info.rows > 0 and empty_rows / info.rows > 0.3:
            score -= 20
            checks.append({
                "dimension": "empty_rows",
                "sheet": info.name,
                "status": "warn",
                "detail": f"Sheet '{info.name}': {empty_rows}/{info.rows} rows are empty ({100*empty_rows//max(info.rows,1)}%)",
            })

        # Check 2: Single column detection
        if info.cols <= 1:
            score -= 15
            checks.append({
                "dimension": "column_count",
                "sheet": info.name,
                "status": "warn",
                "detail": f"Sheet '{info.name}' has only {info.cols} column(s) — may not be a table",
            })

        # Check 3: Very wide tables (possible parsing error)
        if info.cols > 50:
            checks.append({
                "dimension": "column_count",
                "sheet": info.name,
                "status": "info",
                "detail": f"Sheet '{info.name}' has {info.cols} columns — verify column alignment",
            })

        # Check 4: Missing header
        if not info.has_header:
            score -= 10
            checks.append({
                "dimension": "header_detection",
                "sheet": info.name,
                "status": "warn",
                "detail": f"Sheet '{info.name}' has no detectable header row",
            })

        scores.append(max(score, 0))

    # Issues from reading
    if issues:
        checks.append({
            "dimension": "read_issues",
            "status": "warn",
            "detail": f"{len(issues)} issue(s) during reading: {'; '.join(issues[:3])}",
        })

    overall = sum(scores) / max(len(scores), 1)
    if any(c["status"] == "warn" for c in checks):
        overall = min(overall, 85)

    return {
        "overall_score": round(overall, 1),
        "sheets_checked": len(sheet_infos),
        "checks": checks,
    }


def print_xlsx_quality(report: dict[str, Any]) -> None:
    icon = "[PASS]" if report["overall_score"] >= 90 else ("[WARN]" if report["overall_score"] >= 70 else "[FAIL]")
    print(f"""
======================================================================
  XLSX Extraction Quality Report
======================================================================
  Overall:    {icon} {report['overall_score']:.0f}/100
  Sheets:     {report['sheets_checked']} checked
----------------------------------------------------------------------""")
    for c in report["checks"]:
        s = c["status"]
        ic = {"pass": "[OK]  ", "warn": "[WARN]", "info": "[INFO]"}.get(s, "      ")
        print(f"  {ic} {c['detail'][:67]}")
    print("======================================================================\n")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def process_spreadsheet(
    filepath: str,
    output_dir: str,
    *,
    forward_fill: bool = False,
    verify: bool = False,
) -> SpreadsheetManifest:
    """Run full preprocessing on a single spreadsheet file."""
    os.makedirs(output_dir, exist_ok=True)
    source_name = Path(filepath).stem

    # --- Step 1: Read all sheets ---
    sheets, issues = _read_xlsx_sheets(filepath)

    if not sheets and issues:
        # Fatal
        manifest = SpreadsheetManifest(
            source=os.path.basename(filepath),
            total_sheets=0,
            sheets_exported=0,
            total_rows=0,
            total_cells=0,
            issues=issues,
        )
        return manifest

    # --- Step 1.5: Forward-fill merged cells ---
    if forward_fill:
        for sheet_name in list(sheets.keys()):
            before = sheets[sheet_name]
            header_idx = _guess_header_index(before)
            after = _forward_fill(before, start_row=header_idx + 1)
            sheets[sheet_name] = after
            filled_cells = sum(
                1 for old_row, new_row in zip(before, after)
                for old_cell, new_cell in zip(
                    (old_row + [None] * max(0, len(new_row) - len(old_row))),
                    new_row,
                )
                if (old_cell is None or str(old_cell).strip() == "")
                and new_cell is not None
            )
            if filled_cells:
                logger.info(f"Forward-filled {filled_cells} cell(s) in sheet '{sheet_name}'")

    # --- Step 2: Convert each sheet to Markdown ---
    sheets_dir = os.path.join(output_dir, "sheets")
    os.makedirs(sheets_dir, exist_ok=True)

    sheet_infos: list[SheetInfo] = []
    total_rows = 0
    total_cells = 0

    for i, (sheet_name, rows) in enumerate(sheets.items()):
        md, info = _sheet_to_markdown(rows, sheet_name)
        info.index = i + 1

        # Sanitize filename
        safe_name = sheet_name.replace("/", "-").replace("\\", "-")[:50]
        filename = f"sheet-{i + 1:02d}-{safe_name}.md"
        filepath_md = os.path.join(sheets_dir, filename)
        info.markdown_path = filepath_md

        with open(filepath_md, "w", encoding="utf-8") as f:
            f.write(f"<!-- Source: {os.path.basename(filepath)}, sheet: {sheet_name} -->\n\n")
            f.write(md)

        sheet_infos.append(info)
        total_rows += info.rows
        total_cells += info.rows * info.cols

    logger.info(f"Exported {len(sheet_infos)} sheet(s) → {sheets_dir}")

    # --- Step 3: Build manifest ---
    manifest = SpreadsheetManifest(
        source=os.path.basename(filepath),
        total_sheets=len(sheets),
        sheets_exported=len(sheet_infos),
        total_rows=total_rows,
        total_cells=total_cells,
        sheets=sheet_infos,
        issues=issues,
    )

    manifest_path = os.path.join(output_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(asdict(manifest), f, indent=2, ensure_ascii=False)
    logger.info(f"Manifest → {manifest_path}")

    _print_summary(manifest)

    if verify:
        report = verify_xlsx_quality(sheets, sheet_infos, issues)
        report_path = os.path.join(output_dir, "quality-report.json")
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        print_xlsx_quality(report)
        logger.info(f"Quality report -> {report_path}")

    return manifest


def process_batch(input_dir: str, output_base: str, forward_fill: bool = False) -> list[SpreadsheetManifest]:
    """Process all supported spreadsheet files in a directory."""
    results: list[SpreadsheetManifest] = []
    files = [
        f for f in Path(input_dir).iterdir()
        if f.suffix.lower() in SUPPORTED_EXTENSIONS
    ]
    logger.info(f"Batch processing {len(files)} spreadsheet(s) in {input_dir}")

    for f in files:
        output_dir = os.path.join(output_base, f.stem)
        logger.info(f"--- Processing: {f.name} ---")
        manifest = process_spreadsheet(str(f), output_dir, forward_fill=forward_fill)
        results.append(manifest)

    return results


def _print_summary(m: SpreadsheetManifest) -> None:
    print(f"""
═══════════════════════════════════════════
  Spreadsheet Preprocessing Complete
═══════════════════════════════════════════
  Source:          {m.source}
  Sheets:          {m.sheets_exported}/{m.total_sheets} exported
  Total rows:      {m.total_rows}
  Total cells:     {m.total_cells}
  Issues:          {len(m.issues)}
═══════════════════════════════════════════
""")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Spreadsheet Preprocessing for SE Requirements Decomposition",
    )
    parser.add_argument(
        "--input", "-i", required=True,
        help="Path to an .xlsx/.csv/.tsv file or directory (with --batch)",
    )
    parser.add_argument(
        "--output", "-o", required=True,
        help="Output directory for preprocessed artifacts",
    )
    parser.add_argument(
        "--batch", action="store_true",
        help="Process all supported files in the input directory",
    )
    parser.add_argument(
        "--max-rows", type=int, default=5000,
        help="Max rows per sheet (default: 5000)",
    )
    parser.add_argument(
        "--verbose", "-v", action="store_true",
        help="Verbose logging",
    )
    parser.add_argument(
        "--forward-fill", action="store_true",
        help="Forward-fill empty cells from last non-empty value (handles merged cells)",
    )
    parser.add_argument(
        "--verify", action="store_true",
        help="Run quality assessment on extraction results",
    )

    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if args.batch:
        if not os.path.isdir(args.input):
            print(f"ERROR: --batch requires a directory, got: {args.input}", file=sys.stderr)
            sys.exit(1)
        process_batch(args.input, args.output, forward_fill=args.forward_fill)
    else:
        if not os.path.isfile(args.input):
            print(f"ERROR: File not found: {args.input}", file=sys.stderr)
            sys.exit(1)
        process_spreadsheet(args.input, args.output, forward_fill=args.forward_fill, verify=args.verify)


if __name__ == "__main__":
    main()
